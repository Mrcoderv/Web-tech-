from flask import Flask, render_template, request
import openpyxl
import os

app = Flask(__name__)

# Define the path to the Excel file
EXCEL_FILE = "projects.xlsx"

# Create the Excel file if it doesn't exist
if not os.path.exists(EXCEL_FILE):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Projects"
    # Add headers
    sheet.append(["Serial Number", "Creator Name", "Project Name", "Project Link"])
    workbook.save(EXCEL_FILE)

@app.route('/')
def index():
    # Load the data from the Excel file
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Read data into a list of dictionaries
    projects = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip the header row
        projects.append({
            "serial_number": row[0],
            "creator_name": row[1],
            "project_name": row[2],
            "project_link": row[3]
        })

    # Render the template with the project data
    return render_template('index.html', projects=projects)

@app.route('/upload')
def upload_form():
    return render_template('upload_form.html')

@app.route('/submit', methods=['POST'])
def submit():
    creator_name = request.form['creatorName']
    project_name = request.form['projectName']
    project_link = request.form['projectLink']

    # Open the Excel file and append the data
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook.active

    # Get the next serial number
    serial_number = sheet.max_row  # Assuming the first row is the header
    sheet.append([serial_number, creator_name, project_name, project_link])
    workbook.save(EXCEL_FILE)

    return f"<h2>Project successfully uploaded!</h2><p><a href='/'>Go back</a></p>"

if __name__ == '__main__':
    app.run(debug=True)
